#!/usr/bin/env python3
"""
ETF Daily Data Updater - COMPLETE REWRITE
Fetches OHLC data and writes clean, valid prices only
"""

import sys
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import warnings
warnings.filterwarnings('ignore')

# Tickers to fetch (only the ones you need for OHLC)
TICKERS = ['SOXL', 'TQQQ', 'SOXS', 'SQQQ', 'SMH', 'QQQ', 'SOXX']

def fetch_clean_data(ticker, start_date, end_date):
    """Fetch and validate OHLC data"""
    try:
        stock = yf.Ticker(ticker)
        data = stock.history(start=start_date, end=end_date)
        
        if data.empty:
            print(f"  WARNING: No data for {ticker}")
            return None
        
        # Validate prices are positive and reasonable
        for col in ['Open', 'High', 'Low', 'Close']:
            if col in data.columns:
                # Replace any negative or zero prices with NaN
                data.loc[data[col] <= 0, col] = None
        
        # Drop rows where all price data is missing
        data = data.dropna(subset=['Open', 'High', 'Low', 'Close'], how='all')
        
        return data
    except Exception as e:
        print(f"  ERROR fetching {ticker}: {e}")
        return None


def calculate_returns(df):
    """Calculate 1D, 3D, 5D returns from close prices"""
    if df is None or df.empty:
        return pd.DataFrame()
    
    returns_1d = df['Close'].pct_change() * 100
    returns_3d = df['Close'].pct_change(periods=3) * 100
    returns_5d = df['Close'].pct_change(periods=5) * 100
    
    result = pd.DataFrame(index=df.index)
    result['1d_return'] = returns_1d.round(4)
    result['3d_return'] = returns_3d.round(4)
    result['5d_return'] = returns_5d.round(4)
    
    return result


def update_workbook(workbook_path):
    """Main function - fetches data and writes clean Excel file"""
    
    print("=" * 60)
    print("ETF DAILY DATA UPDATER - CLEAN VERSION")
    print("=" * 60)
    
    # Date range: last 400 days
    end_date = datetime.now()
    start_date = end_date - timedelta(days=400)
    
    print(f"\nFetching data from {start_date.date()} to {end_date.date()}")
    
    # Fetch all data
    all_data = {}
    for ticker in TICKERS:
        print(f"Fetching {ticker}...")
        df = fetch_clean_data(ticker, start_date, end_date)
        if df is not None:
            all_data[ticker] = df
    
    if not all_data:
        print("ERROR: No data fetched")
        return False
    
    # Get all dates
    all_dates = set()
    for df in all_data.values():
        all_dates.update(df.index.strftime('%Y-%m-%d'))
    all_dates = sorted(all_dates)
    
    print(f"\nTotal dates: {len(all_dates)}")
    print(f"Date range: {all_dates[0]} to {all_dates[-1]}")
    
    # Build the complete DataFrame
    rows = []
    for date_str in all_dates:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        row = {'Date': date_str}
        
        for ticker, df in all_data.items():
            if date_obj in df.index:
                data = df.loc[date_obj]
                row[f"{ticker}_Open"] = round(data['Open'], 4) if pd.notna(data['Open']) else None
                row[f"{ticker}_High"] = round(data['High'], 4) if pd.notna(data['High']) else None
                row[f"{ticker}_Low"] = round(data['Low'], 4) if pd.notna(data['Low']) else None
                row[f"{ticker}_Close"] = round(data['Close'], 4) if pd.notna(data['Close']) else None
            else:
                row[f"{ticker}_Open"] = None
                row[f"{ticker}_High"] = None
                row[f"{ticker}_Low"] = None
                row[f"{ticker}_Close"] = None
        
        # Calculate returns if we have close prices
        for ticker, df in all_data.items():
            if date_obj in df.index and pd.notna(df.loc[date_obj]['Close']):
                close = df.loc[date_obj]['Close']
                # Need previous closes for returns
                pass
        
        rows.append(row)
    
    df_final = pd.DataFrame(rows)
    
    # Calculate returns for each ticker
    for ticker in TICKERS:
        close_col = f"{ticker}_Close"
        if close_col in df_final.columns:
            # Convert to numeric
            df_final[close_col] = pd.to_numeric(df_final[close_col], errors='coerce')
            
            # Calculate returns
            df_final[f"{ticker}_%Chg"] = df_final[close_col].pct_change() * 100
            df_final[f"{ticker}_3D"] = df_final[close_col].pct_change(periods=3) * 100
            df_final[f"{ticker}_5D"] = df_final[close_col].pct_change(periods=5) * 100
            
            # Round
            df_final[f"{ticker}_%Chg"] = df_final[f"{ticker}_%Chg"].round(4)
            df_final[f"{ticker}_3D"] = df_final[f"{ticker}_3D"].round(4)
            df_final[f"{ticker}_5D"] = df_final[f"{ticker}_5D"].round(4)
    
    print(f"\nFinal DataFrame: {len(df_final)} rows, {len(df_final.columns)} columns")
    
    # Write to Excel
    try:
        # Load or create workbook
        if workbook_path.exists():
            wb = load_workbook(workbook_path)
            print(f"Loaded workbook: {workbook_path}")
        else:
            wb = load_workbook()
            print(f"Created new workbook: {workbook_path}")
        
        # Replace Daily_Data sheet
        if 'Daily_Data' in wb.sheetnames:
            wb.remove(wb['Daily_Data'])
        
        ws = wb.create_sheet('Daily_Data')
        
        # Write headers
        headers = list(df_final.columns)
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Write data
        rows_written = 0
        for row_idx, row in df_final.iterrows():
            row_num = row_idx + 2  # +2 because 1-indexed and header row
            for col_idx, header in enumerate(headers, 1):
                value = row[header]
                if pd.notna(value):
                    ws.cell(row=row_num, column=col_idx, value=value)
            rows_written += 1
        
        print(f"Wrote {rows_written} rows")
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            max_len = 10
            col_letter = get_column_letter(col)
            for row in range(1, min(rows_written + 2, 50)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_len = max(max_len, len(str(cell_value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 15)
        
        # Ensure Signal sheet exists
        if 'Signal' not in wb.sheetnames:
            ws_signal = wb.create_sheet('Signal')
            ws_signal['D23'] = 'SOXL'
            ws_signal['D24'] = 'TQQQ'
            ws_signal['D27'] = datetime.now().strftime('%Y-%m-%d')
            print("Created Signal sheet with defaults")
        else:
            print("Signal sheet preserved")
        
        # Save
        wb.save(workbook_path)
        print(f"\n✅ Successfully updated {workbook_path}")
        print(f"   Rows: {rows_written}")
        print(f"   Columns: {len(headers)}")
        
        return True
        
    except Exception as e:
        print(f"ERROR saving workbook: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    from pathlib import Path
    
    if len(sys.argv) > 1:
        workbook_path = Path(sys.argv[1])
    else:
        workbook_path = Path("4_ETF_Trading_Workbook_Template.xlsx")
    
    success = update_workbook(workbook_path)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
