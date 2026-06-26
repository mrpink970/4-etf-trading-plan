#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path
from typing import Dict, Optional, List, Tuple
import os
import smtplib
from email.message import EmailMessage
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook


WORKBOOK_PATH = Path("4_ETF_Trading_Workbook_Template.xlsx")
POSITIONS_PATH = Path("etf_paper_positions.csv")
TRADE_LOG_PATH = Path("etf_paper_trade_log.csv")
PERFORMANCE_PATH = Path("etf_paper_performance.csv")

# Trading parameters
MAX_TRADES = 1  # Only hold ONE position at a time
TRAILING_STOP_PCT = 0.12
POSITION_SIZE_PCT = 0.95  # Use 95% of account balance per trade
STARTING_BALANCE = 5000.0

# Cash zone threshold - only trade when |score| >= this value
MIN_TRADE_SCORE = 2.0

# NEW: Time stop parameters for choppy markets
TIME_STOP_DAYS = 3
TIME_STOP_MIN_RETURN_PCT = 3.0  # Need at least 3% after 3 days or exit

# NEW: Entry filters to avoid buying at highs
ENTRY_MAX_PCT_FROM_HIGH = 5.0  # Don't buy within 5% of 20-day high
ENTRY_MAX_PREV_2D_RETURN = 3.0  # Don't buy after a 2-day run > 3%

# Ranking weights
RANKING_WEIGHTS = {
    '1d_return': 0.30,
    '3d_return': 0.25,
    '5d_return': 0.20,
    'trend_strength': 0.15,
    'volatility_score': 0.10,
}

# SOXL-only trading (keep data for all 4)
BULL_ETFS = {"SOXL"}  # Only SOXL for bull trades
BEAR_ETFS = {"SOXS"}  # Only SOXS for bear trades
ALL_ETFS = BULL_ETFS | BEAR_ETFS  # Keep data for all

# Score smoothing for 3x ETFs (5-day effective window)
_score_history = {}


def normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip().upper()


def safe_float(value) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except Exception:
        return None


def determine_regime(primary_etf: str) -> str:
    if primary_etf in BULL_ETFS:
        return "bull"
    if primary_etf in BEAR_ETFS:
        return "bear"
    return "neutral"


def calculate_account_balance() -> Tuple[float, float]:
    """Calculate current account balance from closed trades and open positions"""
    cash_balance = STARTING_BALANCE
    total_realized_pl = 0.0
    
    if TRADE_LOG_PATH.exists():
        try:
            trade_log = pd.read_csv(TRADE_LOG_PATH)
            if not trade_log.empty and 'gross_pl' in trade_log.columns:
                total_realized_pl = trade_log['gross_pl'].sum()
                cash_balance = STARTING_BALANCE + total_realized_pl
        except Exception as e:
            print(f"Warning: Could not read trade log: {e}")
    
    unrealized_pl = 0.0
    if POSITIONS_PATH.exists():
        try:
            positions = pd.read_csv(POSITIONS_PATH)
            if not positions.empty:
                for _, row in positions.iterrows():
                    entry_price = safe_float(row.get('entry_price', 0))
                    shares = int(row.get('shares', 0))
                    highest_price = safe_float(row.get('highest_price', 0))
                    if entry_price and shares and highest_price:
                        unrealized_pl += (highest_price - entry_price) * shares
        except Exception as e:
            print(f"Warning: Could not read positions: {e}")
    
    total_equity = cash_balance + unrealized_pl
    return cash_balance, total_equity


def calculate_position_shares(account_balance: float, entry_price: float) -> int:
    """Calculate number of shares based on account balance"""
    if entry_price <= 0:
        return 0
    
    position_value = account_balance * POSITION_SIZE_PCT
    shares = int(position_value / entry_price)
    return max(1, min(shares, 1000))


def read_daily_data_wide(daily_ws) -> pd.DataFrame:
    rows = list(daily_ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Daily_Data sheet is empty.")

    header_idx = None
    for i, row in enumerate(rows):
        values = [str(cell).strip() if cell is not None else "" for cell in row]
        if "Date" in values:
            header_idx = i
            break

    if header_idx is None:
        raise ValueError("Could not find Daily_Data header row.")

    headers = [str(h).strip() if h is not None else "" for h in rows[header_idx]]

    records = []
    for row in rows[header_idx + 1:]:
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        rec = {}
        for i, h in enumerate(headers):
            if h == "":
                continue
            rec[h] = row[i] if i < len(row) else None
        records.append(rec)

    df = pd.DataFrame(records)
    if df.empty:
        raise ValueError("Daily_Data has no usable data rows.")

    if "Date" not in df.columns:
        raise ValueError("Daily_Data missing Date column.")

    df["Date"] = pd.to_datetime(df["Date"], errors='coerce')
    df = df[df["Date"].notna()].copy()
    df = df.sort_values("Date").drop_duplicates(subset=["Date"], keep="last")
    df["Date"] = df["Date"].dt.date

    return df


def extract_latest_prices_and_returns(df: pd.DataFrame) -> Tuple[str, Dict[str, Dict[str, Optional[float]]], Dict[str, Dict[str, Optional[float]]]]:
    """Extract latest prices and returns for ranking"""
    latest_row = df.sort_values("Date").iloc[-1]
    latest_date = str(latest_row["Date"])
    
    prices: Dict[str, Dict[str, Optional[float]]] = {}
    returns: Dict[str, Dict[str, Optional[float]]] = {}
    
    for etf in sorted(ALL_ETFS):
        open_col = f"{etf}_Open"
        high_col = f"{etf}_High"
        low_col = f"{etf}_Low"
        close_col = f"{etf}_Close"
        chg_col = f"{etf}_%Chg"
        three_d_col = f"{etf}_3D"
        five_d_col = f"{etf}_5D"
        
        open_val = latest_row.get(open_col)
        high_val = latest_row.get(high_col)
        low_val = latest_row.get(low_col)
        close_val = latest_row.get(close_col)
        
        prices[etf] = {
            "open": safe_float(open_val),
            "high": safe_float(high_val),
            "low": safe_float(low_val),
            "close": safe_float(close_val),
        }
        
        returns[etf] = {
            "1d": safe_float(latest_row.get(chg_col)),
            "3d": safe_float(latest_row.get(three_d_col)),
            "5d": safe_float(latest_row.get(five_d_col)),
        }

    return latest_date, prices, returns


def calculate_etf_score(returns: Dict[str, Optional[float]]) -> float:
    """Calculate a single ETF's score based on returns"""
    ret_1d = returns.get("1d", 0) or 0
    ret_3d = returns.get("3d", 0) or 0
    ret_5d = returns.get("5d", 0) or 0
    
    score = (ret_1d * RANKING_WEIGHTS['1d_return'] +
             ret_3d * RANKING_WEIGHTS['3d_return'] +
             ret_5d * RANKING_WEIGHTS['5d_return'])
    
    if ret_1d > 0 and ret_3d > 0 and ret_5d > 0:
        score += 5 * RANKING_WEIGHTS['trend_strength']
    elif ret_1d < 0 and ret_3d < 0 and ret_5d < 0:
        score -= 5 * RANKING_WEIGHTS['trend_strength']
    
    volatility = abs(ret_1d - ret_3d) if ret_1d and ret_3d else 0
    volatility_score = max(0, 10 - volatility) * RANKING_WEIGHTS['volatility_score']
    score += volatility_score
    
    return round(score, 4)


def get_smoothed_score(ticker: str, returns: Dict[str, Dict[str, Optional[float]]]) -> float:
    """Calculate smoothed score using exponential weighting (5-day window)"""
    raw_score = calculate_etf_score(returns.get(ticker, {}))
    prev_score = _score_history.get(ticker, raw_score)
    
    alpha = 0.20  # 5-day effective window
    smoothed = prev_score * (1 - alpha) + raw_score * alpha
    
    _score_history[ticker] = smoothed
    return round(smoothed, 4)


def rank_etfs(returns: Dict[str, Dict[str, Optional[float]]], regime: str) -> List[Tuple[str, float]]:
    """Rank ETFs based on momentum scores, filtered by regime"""
    scores = {}
    
    if regime == "bull":
        eligible_etfs = BULL_ETFS  # Only SOXL
    elif regime == "bear":
        eligible_etfs = BEAR_ETFS  # Only SOXS
    else:
        return []  # Neutral regime, no trades
    
    for etf in eligible_etfs:
        etf_returns = returns.get(etf, {})
        score = calculate_etf_score(etf_returns)
        scores[etf] = score
    
    return sorted(scores.items(), key=lambda x: x[1], reverse=True)


def load_workbook_state(path: Path) -> Dict[str, object]:
    """Load state from workbook"""
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")

    wb = load_workbook(path, data_only=True)

    if "Signal" not in wb.sheetnames:
        raise ValueError("Workbook missing Signal sheet.")
    if "Daily_Data" not in wb.sheetnames:
        raise ValueError("Workbook missing Daily_Data sheet.")

    signal_ws = wb["Signal"]
    daily_ws = wb["Daily_Data"]

    primary_etf = normalize_text(signal_ws["D23"].value)
    secondary_etf = normalize_text(signal_ws["D24"].value)

    daily_df = read_daily_data_wide(daily_ws)
    daily_date, prices, returns = extract_latest_prices_and_returns(daily_df)
    asof_date = daily_date

    return {
        "date": asof_date,
        "primary_etf": primary_etf,
        "secondary_etf": secondary_etf,
        "regime": determine_regime(primary_etf),
        "prices": prices,
        "returns": returns,
    }


def load_positions() -> pd.DataFrame:
    cols = [
        "ticker",
        "regime",
        "entry_date",
        "entry_price",
        "shares",
        "highest_price",
        "trailing_stop",
        "rank_score_at_entry",
    ]
    if POSITIONS_PATH.exists():
        df = pd.read_csv(POSITIONS_PATH)
        if df.empty:
            return pd.DataFrame(columns=cols)
        for col in cols:
            if col not in df.columns:
                df[col] = None
        return df
    return pd.DataFrame(columns=cols)


def load_trade_log() -> pd.DataFrame:
    cols = [
        "ticker",
        "regime",
        "entry_date",
        "entry_price",
        "exit_date",
        "exit_price",
        "shares",
        "gross_pl",
        "return_pct",
        "exit_reason",
    ]
    if TRADE_LOG_PATH.exists():
        df = pd.read_csv(TRADE_LOG_PATH)
        if df.empty:
            return pd.DataFrame(columns=cols)
        return df
    return pd.DataFrame(columns=cols)


def save_positions(df: pd.DataFrame) -> None:
    cols = [
        "ticker",
        "regime",
        "entry_date",
        "entry_price",
        "shares",
        "highest_price",
        "trailing_stop",
        "rank_score_at_entry",
    ]
    if df.empty:
        pd.DataFrame(columns=cols).to_csv(POSITIONS_PATH, index=False)
    else:
        for col in cols:
            if col not in df.columns:
                df[col] = None
        df[cols].to_csv(POSITIONS_PATH, index=False)


def save_trade_log(df: pd.DataFrame) -> None:
    cols = [
        "ticker",
        "regime",
        "entry_date",
        "entry_price",
        "exit_date",
        "exit_price",
        "shares",
        "gross_pl",
        "return_pct",
        "exit_reason",
    ]
    if df.empty:
        pd.DataFrame(columns=cols).to_csv(TRADE_LOG_PATH, index=False)
    else:
        df[cols].to_csv(TRADE_LOG_PATH, index=False)


def save_performance(trade_log: pd.DataFrame) -> None:
    cols = [
        "total_trades",
        "win_rate",
        "loss_rate",
        "avg_gain_pct",
        "avg_loss_pct",
        "largest_gain_pct",
        "largest_loss_pct",
        "total_gross_pl",
        "expectancy_pct",
    ]

    if trade_log.empty:
        pd.DataFrame(
            [{
                "total_trades": 0,
                "win_rate": 0.0,
                "loss_rate": 0.0,
                "avg_gain_pct": 0.0,
                "avg_loss_pct": 0.0,
                "largest_gain_pct": 0.0,
                "largest_loss_pct": 0.0,
                "total_gross_pl": 0.0,
                "expectancy_pct": 0.0,
            }]
        )[cols].to_csv(PERFORMANCE_PATH, index=False)
        return

    wins = trade_log[trade_log["gross_pl"] > 0].copy()
    losses = trade_log[trade_log["gross_pl"] < 0].copy()

    total_trades = len(trade_log)
    win_rate = len(wins) / total_trades if total_trades else 0.0
    loss_rate = len(losses) / total_trades if total_trades else 0.0
    avg_gain_pct = wins["return_pct"].mean() if not wins.empty else 0.0
    avg_loss_pct = abs(losses["return_pct"].mean()) if not losses.empty else 0.0
    largest_gain_pct = wins["return_pct"].max() if not wins.empty else 0.0
    largest_loss_pct = losses["return_pct"].min() if not losses.empty else 0.0
    total_gross_pl = trade_log["gross_pl"].sum()
    expectancy_pct = (win_rate * avg_gain_pct) - (loss_rate * avg_loss_pct)

    pd.DataFrame(
        [{
            "total_trades": total_trades,
            "win_rate": round(win_rate, 6),
            "loss_rate": round(loss_rate, 6),
            "avg_gain_pct": round(float(avg_gain_pct), 4),
            "avg_loss_pct": round(float(avg_loss_pct), 4),
            "largest_gain_pct": round(float(largest_gain_pct), 4),
            "largest_loss_pct": round(float(largest_loss_pct), 4),
            "total_gross_pl": round(float(total_gross_pl), 2),
            "expectancy_pct": round(float(expectancy_pct), 4),
        }]
    )[cols].to_csv(PERFORMANCE_PATH, index=False)


def update_trailing_stops(
    positions: pd.DataFrame,
    prices: Dict[str, Dict[str, Optional[float]]],
) -> pd.DataFrame:
    if positions.empty:
        return positions

    out = positions.copy()

    for idx, row in out.iterrows():
        ticker = normalize_text(row["ticker"])
        px = prices.get(ticker, {})
        high_price = px.get("high")
        if high_price is None:
            continue

        current_highest = safe_float(row["highest_price"])
        if current_highest is None or high_price > current_highest:
            current_highest = high_price

        trailing_stop = current_highest * (1 - TRAILING_STOP_PCT)

        out.at[idx, "highest_price"] = round(current_highest, 6)
        out.at[idx, "trailing_stop"] = round(trailing_stop, 6)

    return out


def build_exit_list(
    positions: pd.DataFrame,
    current_regime: str,
    ranked_etfs: List[Tuple[str, float]],
    prices: Dict[str, Dict[str, Optional[float]]],
    returns: Dict[str, Dict[str, Optional[float]]],
    asof_date: str,
) -> list[dict[str, str]]:
    """Build exit list with smoothing + time stop + no rotation"""
    exits: list[dict[str, str]] = []

    if positions.empty:
        return exits

    for _, row in positions.iterrows():
        ticker = normalize_text(row["ticker"])
        held_regime = normalize_text(row["regime"]).lower()
        
        # NEW: Use smoothed score
        current_score = get_smoothed_score(ticker, returns)

        # Exit on regime flip
        if current_regime != "neutral" and held_regime != current_regime:
            exits.append({"ticker": ticker, "reason": "regime_flip"})
            continue

        # NEW: Time stop for choppy markets
        entry_date = pd.to_datetime(row["entry_date"])
        current_date = pd.to_datetime(asof_date)
        hold_days = (current_date - entry_date).days
        
        if hold_days >= TIME_STOP_DAYS:
            entry_price = safe_float(row["entry_price"])
            current_price = prices.get(ticker, {}).get("close")
            if entry_price and current_price:
                return_pct = (current_price / entry_price - 1) * 100
                # If less than TIME_STOP_MIN_RETURN_PCT after TIME_STOP_DAYS, exit
                if return_pct < TIME_STOP_MIN_RETURN_PCT:
                    exits.append({"ticker": ticker, "reason": f"time_stop_{hold_days}d_{return_pct:.1f}pct"})
                    continue

        # Cash zone exit with smoothed score
        if abs(current_score) < MIN_TRADE_SCORE:
            exits.append({"ticker": ticker, "reason": f"cash_zone_score_{current_score:.1f}"})
            continue

        # Trailing stop
        low_price = prices.get(ticker, {}).get("low")
        trailing_stop = safe_float(row["trailing_stop"])
        if low_price is not None and trailing_stop is not None and low_price <= trailing_stop:
            exits.append({"ticker": ticker, "reason": "trailing_stop"})
            continue

        # NO ROTATION LOGIC - REMOVED

    return exits


def apply_exits(
    positions: pd.DataFrame,
    exits: list[dict[str, str]],
    asof_date: str,
    prices: Dict[str, Dict[str, Optional[float]]],
    trade_log: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    if positions.empty or not exits:
        return positions, trade_log

    exit_map = {x["ticker"]: x["reason"] for x in exits}
    keep_rows = []
    new_trades = []

    for _, row in positions.iterrows():
        ticker = normalize_text(row["ticker"])
        if ticker not in exit_map:
            keep_rows.append(row.to_dict())
            continue

        exit_price = prices.get(ticker, {}).get("close")
        if exit_price is None:
            exit_price = prices.get(ticker, {}).get("open")
        
        if exit_price is None:
            keep_rows.append(row.to_dict())
            continue

        entry_price = safe_float(row["entry_price"])
        shares = int(row["shares"])
        gross_pl = (exit_price - entry_price) * shares
        return_pct = ((exit_price / entry_price) - 1) * 100 if entry_price else 0.0

        new_trades.append({
            "ticker": ticker,
            "regime": row["regime"],
            "entry_date": row["entry_date"],
            "entry_price": round(entry_price, 6),
            "exit_date": asof_date,
            "exit_price": round(exit_price, 6),
            "shares": shares,
            "gross_pl": round(gross_pl, 2),
            "return_pct": round(return_pct, 4),
            "exit_reason": exit_map[ticker],
        })

    remaining = pd.DataFrame(keep_rows)
    if remaining.empty:
        remaining = pd.DataFrame(columns=[
            "ticker", "regime", "entry_date", "entry_price",
            "shares", "highest_price", "trailing_stop", "rank_score_at_entry"
        ])

    updated_log = pd.concat([trade_log, pd.DataFrame(new_trades)], ignore_index=True)

    return remaining, updated_log


def build_position_row(
    ticker: str,
    regime: str,
    asof_date: str,
    prices: Dict[str, Dict[str, Optional[float]]],
    account_balance: float,
    rank_score: float,
) -> Optional[dict[str, object]]:
    entry_price = prices.get(ticker, {}).get("open")
    high_price = prices.get(ticker, {}).get("high")

    if entry_price is None:
        return None
    if high_price is None:
        high_price = entry_price

    shares = calculate_position_shares(account_balance, entry_price)
    
    if shares == 0:
        return None

    trailing_stop = high_price * (1 - TRAILING_STOP_PCT)

    return {
        "ticker": ticker,
        "regime": regime,
        "entry_date": asof_date,
        "entry_price": round(entry_price, 6),
        "shares": shares,
        "highest_price": round(high_price, 6),
        "trailing_stop": round(trailing_stop, 6),
        "rank_score_at_entry": round(rank_score, 4),
    }


def check_entry_filters(
    ticker: str,
    prices: Dict[str, Dict[str, Optional[float]]],
    returns: Dict[str, Dict[str, Optional[float]]],
    df: pd.DataFrame,
) -> Tuple[bool, str]:
    """NEW: Apply entry filters to avoid buying at highs"""
    
    # Filter 1: Don't buy if near 20-day high
    # Need to get historical data from df
    close_col = f"{ticker}_Close"
    if close_col in df.columns:
        closes = df[close_col].dropna()
        if len(closes) >= 20:
            recent_high = closes.tail(20).max()
            current_price = closes.iloc[-1]
            pct_from_high = ((recent_high - current_price) / recent_high) * 100
            if pct_from_high < ENTRY_MAX_PCT_FROM_HIGH:
                return False, f"near_20d_high ({pct_from_high:.1f}% from high)"
    
    # Filter 2: Don't buy after a 2-day run
    ret_1d = returns.get(ticker, {}).get("1d", 0) or 0
    ret_2d = returns.get(ticker, {}).get("3d", 0) or 0  # 3D includes last 3 days
    
    # Check 2-day cumulative return (using 1d + 3d approximation)
    two_day_return = ret_1d + (ret_2d - ret_1d)  # Rough estimate
    if two_day_return > ENTRY_MAX_PREV_2D_RETURN:
        return False, f"recent_run ({two_day_return:.1f}% in 2 days)"
    
    return True, "filters_passed"


def apply_entries(
    positions: pd.DataFrame,
    current_regime: str,
    ranked_etfs: List[Tuple[str, float]],
    asof_date: str,
    prices: Dict[str, Dict[str, Optional[float]]],
    returns: Dict[str, Dict[str, Optional[float]]],
    account_balance: float,
    df: pd.DataFrame,  # NEW: Pass full dataframe for historical checks
) -> pd.DataFrame:
    """Apply entries with new filters"""
    if len(positions) >= MAX_TRADES:
        return positions

    if current_regime == "neutral":
        print("Neutral regime — staying in cash")
        return positions

    if not ranked_etfs:
        print("No ranked ETFs available — staying in cash")
        return positions

    # Only trade SOXL in bull regime
    if current_regime == "bull":
        top_etf = "SOXL"
        top_score = ranked_etfs[0][1] if ranked_etfs and ranked_etfs[0][0] == "SOXL" else 0
        
        # Check minimum score
        if abs(top_score) < MIN_TRADE_SCORE:
            print(f"Cash zone: SOXL score {top_score:.2f} (|score| < {MIN_TRADE_SCORE}) — staying in cash")
            return positions
        
        # NEW: Check entry filters
        passed, reason = check_entry_filters(top_etf, prices, returns, df)
        if not passed:
            print(f"Entry filter blocked: {reason} — staying in cash")
            return positions
        
        current = positions.copy()
        if not current.empty and normalize_text(current.iloc[0]["ticker"]) == "SOXL":
            return current
        
        row = build_position_row("SOXL", current_regime, asof_date, prices, account_balance, top_score)
        if row is None:
            return current
        
        current = pd.concat([current, pd.DataFrame([row])], ignore_index=True)
        print(f"Entering SOXL with score {top_score:.2f} (filters passed)")
        return current
    
    return positions


def send_email_summary(asof_date: str, primary_etf: str, secondary_etf: str, 
                       regime: str, positions: pd.DataFrame, trade_log: pd.DataFrame,
                       cash_balance: float, total_equity: float, ranked_etfs: List[Tuple[str, float]],
                       new_entries: list = None, new_exits: list = None) -> None:
    
    mail_username = os.environ.get("MAIL_USERNAME")
    mail_password = os.environ.get("MAIL_PASSWORD")
    mail_to = os.environ.get("MAIL_TO")
    
    if not mail_username or not mail_password or not mail_to:
        print("Email credentials not found. Skipping email notification.")
        return
    
    ranking_text = "\n📊 ETF RANKINGS:\n"
    for i, (etf, score) in enumerate(ranked_etfs[:4], 1):
        cash_zone = " (CASH ZONE)" if abs(score) < MIN_TRADE_SCORE else ""
        ranking_text += f"   {i}. {etf}: {score:.2f}{cash_zone}\n"
    
    total_return = total_equity - STARTING_BALANCE
    total_return_pct = (total_return / STARTING_BALANCE) * 100
    
    body = f"""
═══════════════════════════════════════════════════════════
  ETF PAPER TRADING UPDATE - {asof_date}
═══════════════════════════════════════════════════════════

📊 MARKET REGIME: {regime.upper()}
🎯 SIGNAL: {primary_etf} / {secondary_etf}
💰 CASH BALANCE: ${cash_balance:,.2f}
💵 TOTAL EQUITY: ${total_equity:,.2f}
📈 TOTAL RETURN: {total_return_pct:+.1f}% (${total_return:+,.2f})
{ranking_text}
"""
    
    if new_entries and len(new_entries) > 0:
        body += "\n🟢 NEW ENTRY:\n"
        for entry in new_entries:
            body += f"   • {entry['ticker']}: {entry['shares']} shares @ ${entry['price']:.2f}\n"
            body += f"     Stop: ${entry['stop']:.2f}\n"
    
    if new_exits and len(new_exits) > 0:
        body += "\n🔴 POSITION CLOSED:\n"
        for exit_trade in new_exits:
            pl_symbol = "+" if exit_trade['pl'] >= 0 else ""
            body += f"   • {exit_trade['ticker']}: {exit_trade['return_pct']:.1f}% ({pl_symbol}${exit_trade['pl']:.2f})\n"
            body += f"     Reason: {exit_trade['reason']}\n"
    
    if len(positions) > 0:
        body += "\n📈 CURRENT OPEN POSITION:\n"
        for _, row in positions.iterrows():
            entry_price = safe_float(row["entry_price"])
            shares = int(row["shares"])
            score = safe_float(row.get("rank_score_at_entry", 0)) or 0
            stop = safe_float(row["trailing_stop"])
            high = safe_float(row["highest_price"])
            unrealized = (high - entry_price) * shares if high and entry_price else 0
            body += f"   • {row['ticker']}: {shares} shares @ ${entry_price:.2f}\n"
            body += f"     Current (high): ${high:.2f} | Unrealized: +${unrealized:.2f}\n"
            body += f"     Stop: ${stop:.2f} | Entry Score: {score:.2f}\n"
    else:
        body += "\n📈 CURRENT OPEN POSITION: None (in cash)\n"
    
    if len(trade_log) > 0:
        recent = trade_log.tail(5)
        total_pl = trade_log["gross_pl"].sum() if "gross_pl" in trade_log.columns else 0
        winning_trades = len(trade_log[trade_log["gross_pl"] > 0]) if "gross_pl" in trade_log.columns else 0
        win_rate = (winning_trades / len(trade_log) * 100) if len(trade_log) > 0 else 0
        
        body += f"""
📊 PERFORMANCE SUMMARY:
   • Total Trades: {len(trade_log)}
   • Win Rate: {win_rate:.1f}%
   • Total Realized P&L: ${total_pl:.2f}
"""
    
    body += f"""
═══════════════════════════════════════════════════════════
  📊 VIEW FULL DASHBOARD:
  https://mrpink970.github.io/4-etf-trading-plan/Dashboard.html
═══════════════════════════════════════════════════════════

Generated by GitHub Actions - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
"""
    
    try:
        msg = EmailMessage()
        msg.set_content(body)
        msg["Subject"] = f"ETF Trading Update - {asof_date} ({regime.upper()})"
        msg["From"] = mail_username
        msg["To"] = mail_to
        
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(mail_username, mail_password)
            smtp.send_message(msg)
        
        print(f"Email sent successfully to {mail_to}")
    except Exception as e:
        print(f"Failed to send email: {e}")


def main() -> None:
    print("=" * 50)
    print("ETF PAPER TRADING SYSTEM (SOXL-Only + Chop Protection)")
    print("=" * 50)
    print(f"Cash Zone: |smoothed_score| < {MIN_TRADE_SCORE} → Exit to cash")
    print(f"Time Stop: Exit after {TIME_STOP_DAYS} days if < {TIME_STOP_MIN_RETURN_PCT}%")
    print(f"Entry Filters: Avoid 20-day highs & recent runs")
    print(f"Smoothing: 5-day exponential window")
    
    try:
        state = load_workbook_state(WORKBOOK_PATH)
    except Exception as e:
        print(f"ERROR loading workbook: {e}")
        return
    
    asof_date = state["date"]
    primary_etf = state["primary_etf"]
    secondary_etf = state["secondary_etf"]
    current_regime = state["regime"]
    prices = state["prices"]
    returns = state["returns"]
    
    print(f"\n📅 Processing date: {asof_date} (from Daily_Data)")
    
    # Load full DataFrame for entry filters
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    daily_ws = wb["Daily_Data"]
    df = read_daily_data_wide(daily_ws)
    
    # Rank ETFs for this regime
    ranked_etfs = rank_etfs(returns, current_regime)
    print(f"\n📊 ETF Rankings for {current_regime.upper()} regime:")
    for i, (etf, score) in enumerate(ranked_etfs[:4], 1):
        cash_zone = " (CASH ZONE)" if abs(score) < MIN_TRADE_SCORE else ""
        print(f"   {i}. {etf}: {score:.2f}{cash_zone}")
    
    # Calculate account balance
    cash_balance, total_equity = calculate_account_balance()
    print(f"\n💰 Cash balance: ${cash_balance:,.2f}")
    print(f"💵 Total equity: ${total_equity:,.2f}")
    
    # Load existing data
    old_positions = load_positions()
    old_trade_log = load_trade_log()
    
    print(f"📌 Existing positions: {len(old_positions)}")
    
    new_entries = []
    new_exits = []
    
    # Update trailing stops
    positions = update_trailing_stops(old_positions, prices)
    
    # Build exit list (with all new logic)
    exits = build_exit_list(
        positions=positions,
        current_regime=current_regime,
        ranked_etfs=ranked_etfs,
        prices=prices,
        returns=returns,
        asof_date=asof_date,
    )
    
    # Track exits before applying
    if exits:
        for exit_item in exits:
            ticker = exit_item["ticker"]
            reason = exit_item["reason"]
            pos_row = positions[positions["ticker"] == ticker]
            if not pos_row.empty:
                entry_price = safe_float(pos_row.iloc[0]["entry_price"])
                shares = int(pos_row.iloc[0]["shares"])
                exit_price = prices.get(ticker, {}).get("close")
                if exit_price is None:
                    exit_price = prices.get(ticker, {}).get("open")
                if exit_price:
                    pl = (exit_price - entry_price) * shares
                    return_pct = ((exit_price / entry_price) - 1) * 100
                    new_exits.append({
                        "ticker": ticker,
                        "pl": pl,
                        "return_pct": return_pct,
                        "reason": reason,
                    })
    
    # Apply exits
    positions, trade_log = apply_exits(
        positions=positions,
        exits=exits,
        asof_date=asof_date,
        prices=prices,
        trade_log=old_trade_log,
    )
    
    # Process entries (only if no position)
    if len(positions) == 0 and ranked_etfs:
        top_etf, top_score = ranked_etfs[0]
        if abs(top_score) >= MIN_TRADE_SCORE:
            # Check filters
            passed, reason = check_entry_filters(top_etf, prices, returns, df)
            if passed:
                entry_price = prices.get(top_etf, {}).get("open")
                if entry_price:
                    shares = calculate_position_shares(cash_balance, entry_price)
                    new_entries.append({
                        "ticker": top_etf,
                        "price": entry_price,
                        "shares": shares,
                        "stop": entry_price * (1 - TRAILING_STOP_PCT),
                    })
            else:
                print(f"Entry blocked: {reason}")
        else:
            print(f"Cash zone: top score {top_score:.2f} < {MIN_TRADE_SCORE} — no entry")
    
    positions = apply_entries(
        positions=positions,
        current_regime=current_regime,
        ranked_etfs=ranked_etfs,
        asof_date=asof_date,
        prices=prices,
        returns=returns,
        account_balance=cash_balance,
        df=df,
    )
    
    # Save data
    save_positions(positions)
    save_trade_log(trade_log)
    save_performance(trade_log)
    
    # Send email
    send_email_summary(
        asof_date=asof_date,
        primary_etf=primary_etf,
        secondary_etf=secondary_etf,
        regime=current_regime,
        positions=positions,
        trade_log=trade_log,
        cash_balance=cash_balance,
        total_equity=total_equity,
        ranked_etfs=ranked_etfs,
        new_entries=new_entries,
        new_exits=new_exits,
    )
    
    # Print summary
    print("\n" + "=" * 50)
    print("SUMMARY")
    print("=" * 50)
    print(f"Date: {asof_date}")
    print(f"Cash Balance: ${cash_balance:,.2f}")
    print(f"Total Equity: ${total_equity:,.2f}")
    print(f"Regime: {current_regime}")
    print(f"Open positions: {len(positions)}")
    if len(positions) > 0:
        pos = positions.iloc[0]
        print(f"  • {pos['ticker']}: {int(pos['shares'])} shares @ ${pos['entry_price']:.2f}")
        print(f"    Stop: ${pos['trailing_stop']:.2f}")
    else:
        print("  • In cash (no position)")
    print(f"Closed trades logged: {len(trade_log)}")
    
    if new_entries:
        print(f"\n🟢 New entry: {new_entries[0]['ticker']} ({new_entries[0]['shares']} shares @ ${new_entries[0]['price']:.2f})")
    
    if new_exits:
        print(f"\n🔴 Position closed: {new_exits[0]['ticker']} - {new_exits[0]['return_pct']:.1f}% (${new_exits[0]['pl']:.2f})")
        print(f"    Reason: {new_exits[0]['reason']}")
    
    print("=" * 50)


if __name__ == "__main__":
    main()
